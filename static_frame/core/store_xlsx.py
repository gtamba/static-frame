
import typing as tp

import numpy as np

from static_frame.core.util import DtypesSpecifier

from static_frame.core.util import DTYPE_INT_KIND
from static_frame.core.util import DTYPE_STR_KIND
from static_frame.core.util import DTYPE_NAN_KIND
# from static_frame.core.util import DTYPE_DATETIME_KIND
from static_frame.core.util import DTYPE_BOOL
from static_frame.core.util import COMPLEX_TYPES

from static_frame.core.util import BOOL_TYPES
from static_frame.core.util import NUMERIC_TYPES

from static_frame.core.util import _DT64_S
from static_frame.core.util import _DT64_DAY
from static_frame.core.util import AnyCallable


from static_frame.core.frame import Frame

from static_frame.core.store import Store
from static_frame.core.store_filter import StoreFilter
from static_frame.core.store_filter import STORE_FILTER_DEFAULT

from static_frame.core.index import Index
from static_frame.core.index_base import IndexBase
from static_frame.core.index_hierarchy import IndexHierarchy

from static_frame.core.doc_str import doc_inject

if tp.TYPE_CHECKING:
    from xlsxwriter.worksheet import Worksheet  # pylint: disable=W0611
    from xlsxwriter.workbook import Workbook  # pylint: disable=W0611
    from xlsxwriter.format import Format  # pylint: disable=W0611



class StoreXLSX(Store):

    _EXT: tp.FrozenSet[str] =  frozenset(('.xlsx',))

    # _EXT: str = '.xlsx'

    @staticmethod
    def _dtype_to_writer_attr(
            dtype: np.dtype,
            ) -> tp.Tuple[str, bool]:
        '''
        Return a pair of writer function, Boolean, where Boolean denotes if replacements need be applied.
        '''
        kind = dtype.kind
        if dtype == _DT64_S or dtype == _DT64_DAY:
            return 'write_datetime', True
        elif dtype == DTYPE_BOOL:
            return 'write_boolean', False
        elif kind in DTYPE_STR_KIND:
            return 'write_string', False
        elif kind in DTYPE_INT_KIND:
            return 'write_number', False
        elif kind in DTYPE_NAN_KIND:
            return 'write_number', True
        return 'write', True

    @staticmethod
    def _get_format_or_default(
            workbook: 'Workbook', # do not want module level import o
            format_specifier: tp.Optional[tp.Dict[str, tp.Any]]
            ) -> 'Format':
        if format_specifier:
            return workbook.add_format(format_specifier)
        else:
            f = workbook.add_format()
            f.set_bold()
            return f

    @classmethod
    def _get_writer(cls,
            dtype: np.dtype,
            ws: 'Worksheet'
            ) -> AnyCallable: # find better type
        '''
        Return a writer function of the passed in Worksheet.
        '''
        assert isinstance(dtype, np.dtype)

        import xlsxwriter

        writer_attr, replace_active = cls._dtype_to_writer_attr(dtype)
        writer_native = getattr(ws, writer_attr)

        def writer(
                row: int,
                col: int,
                value: tp.Any,
                cell_format: tp.Optional[xlsxwriter.format.Format] = None
                ) -> tp.Any:

            # cannot yet write complex types directly, so covert to string
            if isinstance(value, COMPLEX_TYPES):
                return ws.write_string(row, col, str(value), cell_format)

            if writer_attr == 'write':
                # determine type for aach value
                if isinstance(value, BOOL_TYPES):
                    return ws.write_boolean(row, col, value, cell_format)
                if isinstance(value, str):
                    return ws.write_string(row, col, value, cell_format)
                if isinstance(value, NUMERIC_TYPES):
                    return ws.write_number(row, col, value, cell_format)

            # use the type specific writer_native
            return writer_native(row, col, value, cell_format)

        return writer


    @classmethod
    def _frame_to_worksheet(cls,
            frame: Frame,
            ws: 'Worksheet',
            *,
            include_columns: bool,
            include_index: bool,
            format_columns: 'Format',
            format_index: 'Format',
            merge_hierarchical_labels: bool,
            store_filter: tp.Optional[StoreFilter]
            ) -> None:

        index_depth = frame._index.depth
        index_depth_effective = 0 if not include_index else index_depth

        columns_iter = cls.get_column_iterator(frame=frame,
                include_index=include_index)

        if include_columns:
            columns_depth = frame._columns.depth
            columns_values = frame._columns.values
            if store_filter:
                columns_values = store_filter.from_type_filter_array(columns_values)
            writer_columns = cls._get_writer(columns_values.dtype, ws)



        # TODO: need to determine if .name attr on index or columns should be populated in upper left corner "dead" zone.

        # can write by column
        for col, values in enumerate(columns_iter):

            if include_columns:
                # The col integers will include index depth, so if including index, must wait until after index depth to write column field names; if include_index is False, can begin reading from columns_values
                if col >= index_depth_effective:
                    if columns_depth == 1:
                        writer_columns(0,
                                col,
                                columns_values[col - index_depth_effective],
                                format_columns)
                    elif columns_depth > 1:
                        for i in range(columns_depth):
                            # here, row selection is column count, column selection is depth
                            writer_columns(i,
                                    col,
                                    columns_values[col - index_depth_effective, i],
                                    format_columns
                                    )

            if store_filter:
                # thi might change the dtype
                values = store_filter.from_type_filter_array(values)
            writer = cls._get_writer(values.dtype, ws)
            # start enumeration of row after the columns
            for row, v in enumerate(values, columns_depth):
                writer(row,
                        col,
                        v,
                        format_index if col < index_depth_effective else None)

        # post process to merge cells; need to get width of at depth
        if include_columns and merge_hierarchical_labels and columns_depth > 1:
            for depth in range(columns_depth - 1): # never most deep
                row = depth
                col = index_depth_effective # start after index
                for label, width in frame._columns.label_widths_at_depth(depth):
                    # TODO: use store_filter
                    ws.merge_range(row, col, row, col + width - 1, label, format_columns)
                    col += width

        if include_index and merge_hierarchical_labels and index_depth > 1:
            for depth in range(index_depth - 1): # never most deep
                row = columns_depth
                col = depth
                for label, width in frame._index.label_widths_at_depth(depth):
                    # TODO: use store_filter
                    ws.merge_range(row, col, row + width - 1, col, label, format_columns)
                    row += width


    def write(self,
            items: tp.Iterable[tp.Tuple[tp.Optional[str], Frame]],
            *,
            include_index: bool = True,
            include_columns: bool = True,
            format_index: tp.Optional[tp.Dict[str, tp.Any]] = None,
            format_columns: tp.Optional[tp.Dict[str, tp.Any]] = None,
            merge_hierarchical_labels: bool = True,
            store_filter: tp.Optional[StoreFilter] = STORE_FILTER_DEFAULT
            ) -> None:
        '''
        Args:
            include_index: Boolean to determine if the ``index`` is included in output.
            include_columns: Boolean to determine if the ``columns`` is included in output.
            format_index: dictionary of XlsxWriter format specfications.
            format_columns: dictionary of XlsxWriter format specfications.
            store_filter: a dictionary of objects to string, enabling replacement of NaN and None values when writng to XLSX.
        '''
        # format_data: tp.Optional[tp.Dict[tp.Hashable, tp.Dict[str, tp.Any]]]
        # format_data: dictionary of dictionaries, keyed by column label, that contains dictionaries of XlsxWriter format specifications.

        import xlsxwriter

        wb = xlsxwriter.Workbook(self._fp)

        format_columns = self._get_format_or_default(wb, format_columns)
        format_index = self._get_format_or_default(wb, format_index)

        for label, frame in items:
            ws = wb.add_worksheet(label)
            self._frame_to_worksheet(frame,
                    ws,
                    format_columns=format_columns,
                    format_index=format_index,
                    include_index=include_index,
                    include_columns=include_columns,
                    merge_hierarchical_labels=merge_hierarchical_labels,
                    store_filter=store_filter
                    )

        wb.close()


    @staticmethod
    def _load_workbook(fp: str) -> 'Workbook':
        import openpyxl
        return openpyxl.load_workbook(
                filename=fp,
                read_only=True,
                data_only=True
                )

    @doc_inject(selector='constructor_frame')
    def read(self,
            label: tp.Optional[str] = None,
            *,
            index_depth: int=1,
            columns_depth: int=1,
            dtypes: DtypesSpecifier = None,
            store_filter: tp.Optional[StoreFilter] = STORE_FILTER_DEFAULT
            ) -> Frame:
        '''
        Args:
            {dtypes}
        '''
        wb = self._load_workbook(self._fp)

        if label is None:
            ws = wb[wb.sheetnames[0]]
            name = None # do not set to default sheet name
        else:
            ws = wb[label]
            name = ws.title


        if ws.max_column <= 1 or ws.max_row <= 1:
            # https://openpyxl.readthedocs.io/en/stable/optimized.html
            # says that some clients might not repare correct dimensions; not sure what conditions are best to show this
            ws.calculate_dimension()

        max_column = ws.max_column
        max_row = ws.max_row

        index_values: tp.List[tp.Any] = []
        columns_values: tp.List[tp.Any] = []

        # print()
        # for row in ws.iter_rows():
        #     print(tuple(str(c.value).ljust(10) for c in row))

        data = []

        for row_count, row in enumerate(ws.iter_rows()): # cannot use values_only on 2.5.4
            if store_filter is None:
                row = tuple(c.value for c in row)
            else: # only need to filter string values, but probably too expensive to pre-check
                row = tuple(store_filter.to_type_filter_element(c.value) for c in row)

            if row_count <= columns_depth - 1:
                if columns_depth == 1:
                    columns_values.extend(row[index_depth:])
                elif columns_depth > 1:
                    # NOTE: this orientation will need to be rotated
                    columns_values.append(row[index_depth:])
                continue

            if index_depth == 0:
                data.append(row)
            elif index_depth == 1:
                index_values.append(row[0])
                data.append(row[1:])
            else:
                index_values.append(row[:index_depth])
                data.append(row[index_depth:])

        wb.close()


        index: tp.Optional[IndexBase] = None
        own_index = False
        if index_depth == 1:
            index = Index(index_values)
            own_index = True
        elif index_depth > 1:
            index = IndexHierarchy.from_labels(
                    index_values,
                    continuation_token=None
                    )
            own_index = True

        columns: tp.Optional[IndexBase] = None
        own_columns = False
        if columns_depth == 1:
            columns = Index(columns_values)
            own_columns = True
        elif columns_depth > 1:
            columns = IndexHierarchy.from_labels(
                    zip(*columns_values),
                    continuation_token=None
                    )
            own_columns = True

        return tp.cast(Frame, Frame.from_records(data,
                index=index,
                columns=columns,
                dtypes=dtypes,
                own_index=own_index,
                own_columns=own_columns,
                name=name
                ))

    def labels(self) -> tp.Iterator[str]:
        wb = self._load_workbook(self._fp)
        labels = tuple(wb.sheetnames)
        wb.close()
        yield from labels


# p q I I II II
# r s A B A  B
# 1 A 0 0 0  0
# 1 B 0 0 0  0
